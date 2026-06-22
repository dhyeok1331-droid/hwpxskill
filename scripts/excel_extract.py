#!/usr/bin/env python3
"""Convert Excel workbooks (.xlsx/.xls) to Markdown tables, one section per sheet.

Merged cells are filled with their top-left value so the resulting table
reads naturally (the same value repeats across the merged range) rather
than leaving blank cells that break row alignment.

Usage:
    python excel_extract.py document.xlsx
    python excel_extract.py ./excel_files --recursive --output-dir ./md_out
"""

import argparse
import datetime
import re
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}
_COL_RE = re.compile(r"([A-Z]+)(\d+)")


def _col_to_num(letters: str) -> int:
    num = 0
    for ch in letters:
        num = num * 26 + (ord(ch) - ord("A") + 1)
    return num


def _parse_ref(ref: str) -> tuple[int, int]:
    m = _COL_RE.match(ref)
    return int(m.group(2)), _col_to_num(m.group(1))


def _xlsx_merge_ranges(path: str, sheet_name: str) -> list[tuple[int, int, int, int]]:
    """Read merged-cell ranges for one sheet directly from the raw XML.

    Avoids loading the sheet through openpyxl's object model, since that's
    what made huge sheets (hundreds of thousands of styled cells) hang.
    """
    with zipfile.ZipFile(path) as z:
        wb_xml = ET.fromstring(z.read("xl/workbook.xml"))
        rels_xml = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))

        rid_to_target = {
            el.get("Id"): el.get("Target")
            for el in rels_xml
        }
        target = None
        for sheet_el in wb_xml.find("main:sheets", _NS):
            if sheet_el.get("name") == sheet_name:
                rid = sheet_el.get(f"{{{_NS['r']}}}id")
                target = rid_to_target.get(rid)
                break
        if target is None:
            return []

        sheet_path = "xl/" + target if not target.startswith("/xl/") else target.lstrip("/")
        data = z.read(sheet_path)

    ranges = []
    for m in re.finditer(rb'<mergeCell ref="([^"]+)"', data):
        ref = m.group(1).decode("ascii")
        parts = ref.split(":")
        r1, c1 = _parse_ref(parts[0])
        r2, c2 = _parse_ref(parts[-1])
        ranges.append((r1, c1, r2, c2))
    return ranges


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value)
    return text.replace("\n", "<br>").replace("|", "\\|").strip()


def _grid_to_markdown(grid: list[list[str]]) -> str:
    # drop fully empty trailing rows/columns
    while grid and all(c == "" for c in grid[-1]):
        grid.pop()
    while grid and all(row[-1] == "" for row in grid):
        for row in grid:
            row.pop()
    if not grid:
        return ""

    width = max(len(row) for row in grid)
    for row in grid:
        row.extend([""] * (width - len(row)))

    lines = []
    header, *rest = grid
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join(["---"] * width) + " |")
    for row in rest:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)


MAX_ROWS = 2000
MAX_COLS = 200


def _read_xlsx_sheet(ws, merge_ranges: list[tuple[int, int, int, int]]) -> tuple[list[list[str]], bool]:
    max_row = min(ws.max_row, MAX_ROWS)
    max_col = min(ws.max_column, MAX_COLS)
    truncated = ws.max_row > MAX_ROWS or ws.max_column > MAX_COLS
    grid = [["" for _ in range(max_col)] for _ in range(max_row)]
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col)):
        for c, cell in enumerate(row):
            grid[r][c] = _cell_text(cell.value)
    for r1, c1, r2, c2 in merge_ranges:
        if r1 > max_row or c1 > max_col:
            continue
        fill_value = grid[r1 - 1][c1 - 1]
        for r in range(r1, min(r2, max_row) + 1):
            for c in range(c1, min(c2, max_col) + 1):
                grid[r - 1][c - 1] = fill_value
    return grid, truncated


def _read_xls_sheet(sheet) -> tuple[list[list[str]], bool]:
    max_row = min(sheet.nrows, MAX_ROWS)
    max_col = min(sheet.ncols, MAX_COLS)
    truncated = sheet.nrows > MAX_ROWS or sheet.ncols > MAX_COLS
    grid = [[_cell_text(sheet.cell_value(r, c)) for c in range(max_col)] for r in range(max_row)]
    for crange in sheet.merged_cells:
        rlo, rhi, clo, chi = crange
        if rlo >= max_row or clo >= max_col:
            continue
        fill_value = grid[rlo][clo]
        for r in range(rlo, min(rhi, max_row)):
            for c in range(clo, min(chi, max_col)):
                grid[r][c] = fill_value
    return grid, truncated


def extract_markdown(path: str) -> str:
    """Convert an Excel workbook to Markdown, one section per sheet."""

    suffix = Path(path).suffix.lower()
    sections = []

    if suffix == ".xlsx":
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        for name in wb.sheetnames:
            merge_ranges = _xlsx_merge_ranges(path, name)
            grid, truncated = _read_xlsx_sheet(wb[name], merge_ranges)
            table = _grid_to_markdown(grid)
            note = f"\n\n*(시트가 {MAX_ROWS}행x{MAX_COLS}열을 넘어 앞부분만 표시됨. 전체 데이터는 원본 파일 참고)*" if truncated else ""
            sections.append(f"## {name}\n\n{table}{note}" if table else f"## {name}\n\n(빈 시트)")
        wb.close()
    elif suffix == ".xls":
        import xlrd

        wb = xlrd.open_workbook(path)
        for sheet in wb.sheets():
            grid, truncated = _read_xls_sheet(sheet)
            table = _grid_to_markdown(grid)
            note = f"\n\n*(시트가 {MAX_ROWS}행x{MAX_COLS}열을 넘어 앞부분만 표시됨. 전체 데이터는 원본 파일 참고)*" if truncated else ""
            sections.append(f"## {sheet.name}\n\n{table}{note}" if table else f"## {sheet.name}\n\n(빈 시트)")
    else:
        raise ValueError(f"Unsupported Excel format: {suffix}")

    return "\n\n---\n\n".join(sections)


TIMEOUT_SECONDS = 90


def _convert_one(src: str, dest: str) -> None:
    """Run in a separate process so the parent can enforce a timeout."""
    result = extract_markdown(src)
    Path(dest).write_text(result, encoding="utf-8")


def convert_folder(input_dir: Path, *, output_dir: Path | None, recursive: bool, skip_existing: bool) -> tuple[int, int, int]:
    import multiprocessing

    pattern = "**/*" if recursive else "*"
    files = sorted(
        p for p in input_dir.glob(pattern)
        if p.is_file() and p.suffix.lower() in (".xlsx", ".xls")
    )
    if not files:
        print(f"No .xlsx/.xls files found in: {input_dir}", file=sys.stderr)
        sys.exit(1)

    ok, failed, skipped = 0, 0, 0
    for src in files:
        if output_dir:
            dest_dir = output_dir / src.parent.relative_to(input_dir)
            dest_dir.mkdir(parents=True, exist_ok=True)
        else:
            dest_dir = src.parent
        dest = dest_dir / (src.stem + ".md")

        if skip_existing and dest.exists() and dest.stat().st_mtime >= src.stat().st_mtime:
            print(f"SKIP {src} (up to date)", flush=True)
            skipped += 1
            continue

        proc = multiprocessing.Process(target=_convert_one, args=(str(src), str(dest)))
        proc.start()
        proc.join(TIMEOUT_SECONDS)
        if proc.is_alive():
            proc.terminate()
            proc.join(5)
            print(f"FAIL {src}: timed out after {TIMEOUT_SECONDS}s, skipped", file=sys.stderr, flush=True)
            failed += 1
        elif proc.exitcode == 0 and dest.exists():
            print(f"OK   {src} -> {dest}", flush=True)
            ok += 1
        else:
            print(f"FAIL {src}: worker process exited with code {proc.exitcode}", file=sys.stderr, flush=True)
            failed += 1

    return ok, failed, skipped


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert Excel workbooks to Markdown tables"
    )
    parser.add_argument("input", help="Path to .xlsx/.xls file, or a folder")
    parser.add_argument(
        "--output-dir", "-o",
        help="Folder to write .md files to (default: same folder as each source)",
    )
    parser.add_argument(
        "--recursive", "-r",
        action="store_true",
        help="Search subfolders as well (folder mode only)",
    )
    parser.add_argument(
        "--skip-existing",
        action="store_true",
        help="Skip files that already have a matching .md output (for resuming an interrupted run)",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    output_dir = Path(args.output_dir) if args.output_dir else None

    if input_path.is_file():
        result = extract_markdown(str(input_path))
        if output_dir:
            output_dir.mkdir(parents=True, exist_ok=True)
            dest = output_dir / (input_path.stem + ".md")
        else:
            dest = input_path.with_suffix(".md")
        dest.write_text(result, encoding="utf-8")
        print(f"Converted: {dest}")
        return

    if not input_path.is_dir():
        print(f"Error: Path not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    if output_dir:
        output_dir.mkdir(parents=True, exist_ok=True)

    ok, failed, skipped = convert_folder(
        input_path, output_dir=output_dir, recursive=args.recursive, skip_existing=args.skip_existing
    )
    print(f"\nDone: {ok} converted, {skipped} skipped, {failed} failed.", file=sys.stderr)
    if failed:
        sys.exit(1)


if __name__ == "__main__":
    main()
